"""
Excel 结构化解析器
使用 excel-to-markdown 库进行更准确的表头检测和 Markdown 转换
"""

import os
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional
import logging

import pandas as pd
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# 尝试导入 excel-to-markdown 库
try:
    from excel_to_markdown.detector import get_table_region, detect_table_start
    from excel_to_markdown.markdown_generator import dataframe_to_markdown as excel_lib_markdown
    EXCEL_TO_MARKDOWN_AVAILABLE = True
    logger.info("✅ excel-to-markdown 库已加载")
except ImportError as e:
    EXCEL_TO_MARKDOWN_AVAILABLE = False
    logger.warning(f"⚠️ excel-to-markdown 库不可用: {e}，将使用备用方案")

# openpyxl 始终导入（用于类型注解和备用方案）
try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    load_workbook = None
    Worksheet = None  # type: ignore


def _detect_header_rows(ws: "Worksheet", max_header_rows: int = 5) -> int:
    """粗略检测表头占用的行数（1-5 行）。
    策略：
    - 前几行文本占比更高，且非空单元格较集中
    - 出现明显数值列则认为数据区开始
    """
    header_rows = 1
    for i in range(1, min(ws.max_row, max_header_rows) + 1):
        row_values = [cell.value for cell in ws[i]]
        non_empty = [v for v in row_values if v is not None and str(v).strip() != ""]
        if not non_empty:
            continue
        num_like = sum(isinstance(v, (int, float)) for v in non_empty)
        text_like = len(non_empty) - num_like
        # 若文本明显多于数字，可能仍在表头
        if text_like >= num_like:
            header_rows = i
        else:
            break
    return max(1, header_rows)


def _merged_header(ws: "Worksheet", header_rows: int) -> List[List[str]]:
    """构建多级表头，展开合并单元格并向下填充。"""
    grid: List[List[str]] = []
    for r in range(1, header_rows + 1):
        row_vals: List[str] = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            val = str(cell.value).strip() if cell.value is not None else ""
            row_vals.append(val)
        grid.append(row_vals)

    # 向右填充空表头（继承左侧值），向下填充（继承上一层）
    for r in range(len(grid)):
        last = ""
        for c in range(len(grid[r])):
            if grid[r][c] == "":
                grid[r][c] = last
            last = grid[r][c]

    for r in range(1, len(grid)):
        for c in range(len(grid[r])):
            if grid[r][c] == "":
                grid[r][c] = grid[r - 1][c]

    return grid


def _headers_to_single_row(headers_grid: List[List[str]]) -> List[str]:
    """将多级表头压平成单行列名，以 ' / ' 连接。"""
    levels = len(headers_grid)
    cols = len(headers_grid[0]) if headers_grid else 0
    single: List[str] = []
    for c in range(cols):
        parts = []
        for r in range(levels):
            val = headers_grid[r][c].strip()
            if val:
                parts.append(val)
        single.append(" / ".join(parts) if parts else f"col_{c+1}")
    return single


def _infer_column_types(df: pd.DataFrame) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    for col in df.columns:
        dtype = str(df[col].dtype)
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            mapping[col] = "datetime"
        elif pd.api.types.is_integer_dtype(df[col]):
            mapping[col] = "integer"
        elif pd.api.types.is_float_dtype(df[col]):
            mapping[col] = "float"
        elif pd.api.types.is_bool_dtype(df[col]):
            mapping[col] = "bool"
        else:
            mapping[col] = "text"
    return mapping


def _df_to_markdown(df: pd.DataFrame, max_rows: int = 2000) -> str:
    """将 DataFrame 转换为 Markdown 表格"""
    if df.empty:
        return ""
    
    preview = df.head(max_rows)
    
    # 优先使用 excel-to-markdown 库的转换函数（更准确）
    if EXCEL_TO_MARKDOWN_AVAILABLE:
        try:
            return excel_lib_markdown(preview)
        except Exception as e:
            logger.warning(f"excel-to-markdown 转换失败，使用 pandas fallback: {e}")
    
    # 备用方案：使用 pandas 的 to_markdown（需要 tabulate）
    try:
        return preview.to_markdown(index=False)
    except ImportError:
        # 如果 tabulate 不可用，手动生成 Markdown
        return _manual_markdown_table(preview)


def _manual_markdown_table(df: pd.DataFrame) -> str:
    """手动生成 Markdown 表格（当 tabulate 不可用时）"""
    if df.empty:
        return ""
    
    # 表头
    header = "| " + " | ".join(str(col) for col in df.columns) + " |\n"
    separator = "| " + " | ".join(["---"] * len(df.columns)) + " |\n"
    
    # 数据行
    rows = []
    for _, row in df.iterrows():
        row_values = [str(cell) if pd.notnull(cell) else "" for cell in row]
        rows.append("| " + " | ".join(row_values) + " |\n")
    
    return header + separator + "".join(rows)


def parse_excel_to_markdown(file_path: str) -> Dict[str, Any]:
    """解析 Excel，使用 excel-to-markdown 库进行更准确的表头检测和转换。

    Returns:
        {
          'file_type': 'excel',
          'sheets': [ {sheet_name, headers_levels, headers_flat, types, rows, markdown} ],
          'content': '拼接后的markdown',
          'metadata': {...}
        }
    """
    logger.info(f"🔄 开始处理 Excel 文件: {file_path}")
    if not os.path.exists(file_path):
        raise FileNotFoundError(file_path)

    result_sheets: List[Dict[str, Any]] = []
    md_parts: List[str] = []

    # 使用 excel-to-markdown 库（更好的表头检测）
    if EXCEL_TO_MARKDOWN_AVAILABLE:
        try:
            # 获取所有工作表
            excel_file = pd.ExcelFile(file_path, engine='openpyxl')
            sheet_names = excel_file.sheet_names
            
            for sheet_name in sheet_names:
                logger.info(f"📊 处理工作表: {sheet_name}")
                try:
                    # 读取整个工作表（无表头）
                    df_full = pd.read_excel(
                        file_path, 
                        sheet_name=sheet_name, 
                        header=None, 
                        engine='openpyxl'
                    )
                    
                    if df_full.empty or df_full.shape[0] == 0:
                        logger.warning(f"⚠️ 工作表 '{sheet_name}' 为空，跳过")
                        continue
                    
                    # 使用 excel-to-markdown 的自动检测（非交互式）
                    try:
                        # 直接使用 detect_table_start，避免交互式输入
                        start_row = detect_table_start(df_full)
                        if start_row is not None:
                            headers_row = int(start_row)
                            # 检测有效列（至少有49%非空值）
                            threshold = 0.49
                            valid_col_indices = [
                                col_idx for col_idx in df_full.columns 
                                if df_full[col_idx].notnull().mean() > threshold
                            ]
                            # 转换为列索引列表（0-based）或列名
                            usecols = valid_col_indices if valid_col_indices else None
                            logger.info(f"✅ 自动检测到表头行: {headers_row + 1}, 有效列: {len(usecols) if usecols else '全部'}")
                        else:
                            logger.warning(f"⚠️ 自动检测失败，使用默认策略（第一行作为表头）")
                            headers_row = 0
                            usecols = None
                    except Exception as e:
                        logger.warning(f"⚠️ 自动检测失败: {e}，使用默认策略（第一行作为表头）")
                        headers_row = 0
                        usecols = None
                    
                    # 读取数据（使用检测到的参数）
                    # 注意：header 参数需要使用检测到的行作为表头
                    df = pd.read_excel(
                        file_path,
                        sheet_name=sheet_name,
                        header=headers_row if headers_row is not None else 0,
                        usecols=usecols,
                        engine='openpyxl'
                    )
                    
                    # 清理空行
                    df.dropna(how='all', inplace=True)
                    df.reset_index(drop=True, inplace=True)
                    
                    if df.empty:
                        logger.warning(f"⚠️ 工作表 '{sheet_name}' 处理后为空，跳过")
                        continue
                    
                    # 推断列类型
                    types = _infer_column_types(df)
                    
                    # 转换为 Markdown
                    md = _df_to_markdown(df) if not df.empty else ""
                    
                    # 生成工作表段落
                    section_md = f"### 工作表: {sheet_name}\n\n{md}\n"
                    md_parts.append(section_md)
                    
                    result_sheets.append({
                        "sheet_name": sheet_name,
                        "headers_levels": [[str(col) for col in df.columns]],  # 简化格式
                        "headers_flat": [str(col) for col in df.columns],
                        "data_types": types,
                        "row_count": int(df.shape[0]),
                        "markdown": md,
                    })
                    logger.info(f"✅ 工作表 '{sheet_name}' 处理完成: {df.shape[0]} 行, {df.shape[1]} 列")
                    
                except Exception as e:
                    logger.error(f"❌ 处理工作表 '{sheet_name}' 失败: {e}", exc_info=True)
                    continue
            
            excel_file.close()
            
        except Exception as e:
            logger.error(f"❌ excel-to-markdown 库处理失败: {e}，回退到备用方案", exc_info=True)
            # 回退到旧方案
            return _parse_excel_fallback(file_path)
    else:
        # 使用备用方案
        return _parse_excel_fallback(file_path)

    final_md = "\n\n".join(md_parts)

    # 保存中间 Markdown 以便调试（默认启用）
    # try:
    #     debug_flag = str(os.getenv("EXCEL_DEBUG_SAVE", "1")).lower() in ("1", "true", "yes", "on")
    #     if debug_flag:
    #         debug_root = os.getenv("EXCEL_DEBUG_DIR", "debug_outputs/excel")
    #         debug_dir = Path(debug_root)
    #         debug_dir.mkdir(parents=True, exist_ok=True)

    #         src = Path(file_path)
    #         # 使用文件名（不含路径）作为基础
    #         base = src.stem
    #         timestamp = None
    #         try:
    #             from datetime import datetime
    #             timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    #         except:
    #             pass
            
    #         # 保存整合后的 Markdown
    #         combined_filename = f"{base}__combined.md"
    #         if timestamp:
    #             combined_filename = f"{base}__{timestamp}__combined.md"
    #         combined_path = debug_dir / combined_filename
    #         combined_path.write_text(final_md, encoding="utf-8")
    #         logger.info(f"💾 保存合并 Markdown: {combined_path}")

    #         # 分 sheet 保存
    #         for sheet in result_sheets:
    #             sheet_md = sheet.get("markdown", "")
    #             sheet_name = sheet.get("sheet_name", "sheet")
    #             safe_sheet = "".join(ch if ch.isalnum() or ch in ("_", "-", ".") else "_" for ch in sheet_name)
    #             sheet_filename = f"{base}__{safe_sheet}.md"
    #             if timestamp:
    #                 sheet_filename = f"{base}__{timestamp}__{safe_sheet}.md"
    #             sheet_path = debug_dir / sheet_filename
    #             sheet_path.write_text(f"### 工作表: {sheet_name}\n\n{sheet_md}\n", encoding="utf-8")
    #             logger.debug(f"💾 保存工作表 Markdown: {sheet_path}")
            
    #         logger.info(f"✅ Excel 调试 Markdown 已保存至: {debug_dir} (共 {len(result_sheets) + 1} 个文件)")
    # except Exception as e:
    #     logger.warning(f"⚠️ Excel 调试 Markdown 保存失败: {e}", exc_info=True)

    return {
        "file_type": "excel",
        "sheets": result_sheets,
        "content": final_md,
        "metadata": {
            "sheet_count": len(result_sheets),
            "file_size": os.path.getsize(file_path)
        }
    }


def _parse_excel_fallback(file_path: str) -> Dict[str, Any]:
    """备用解析方案（当 excel-to-markdown 不可用时）"""
    if load_workbook is None:
        raise ImportError("openpyxl 不可用，无法使用备用解析方案")
    
    logger.info("使用备用 Excel 解析方案")
    wb = load_workbook(file_path, read_only=True, data_only=True)
    result_sheets: List[Dict[str, Any]] = []
    md_parts: List[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row == 0 or ws.max_column == 0:
            continue

        header_rows = _detect_header_rows_fallback(ws)
        header_grid = _merged_header_fallback(ws, header_rows)
        headers_flat = _headers_to_single_row(header_grid)

        # 读取数据区
        data: List[List[Any]] = []
        for r in range(header_rows + 1, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            if any(v is not None and str(v).strip() != "" for v in row_vals):
                data.append(row_vals)

        df = pd.DataFrame(data, columns=headers_flat if headers_flat else None)
        types = _infer_column_types(df) if not df.empty else {}
        md = _df_to_markdown(df) if not df.empty else ""

        section_md = f"### 工作表: {sheet_name}\n\n{md}\n"
        md_parts.append(section_md)

        result_sheets.append({
            "sheet_name": sheet_name,
            "headers_levels": header_grid,
            "headers_flat": headers_flat,
            "data_types": types,
            "row_count": int(df.shape[0]) if not df.empty else 0,
            "markdown": md,
        })

    wb.close()
    final_md = "\n\n".join(md_parts)

    return {
        "file_type": "excel",
        "sheets": result_sheets,
        "content": final_md,
        "metadata": {
            "sheet_count": len(result_sheets),
            "file_size": os.path.getsize(file_path)
        }
    }


def _detect_header_rows_fallback(ws: "Worksheet", max_header_rows: int = 5) -> int:
    """粗略检测表头占用的行数（备用方案）"""
    header_rows = 1
    for i in range(1, min(ws.max_row, max_header_rows) + 1):
        row_values = [cell.value for cell in ws[i]]
        non_empty = [v for v in row_values if v is not None and str(v).strip() != ""]
        if not non_empty:
            continue
        num_like = sum(isinstance(v, (int, float)) for v in non_empty)
        text_like = len(non_empty) - num_like
        if text_like >= num_like:
            header_rows = i
        else:
            break
    return max(1, header_rows)


def _merged_header_fallback(ws: "Worksheet", header_rows: int) -> List[List[str]]:
    """构建多级表头（备用方案）"""
    grid: List[List[str]] = []
    for r in range(1, header_rows + 1):
        row_vals: List[str] = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            val = str(cell.value).strip() if cell.value is not None else ""
            row_vals.append(val)
        grid.append(row_vals)

    for r in range(len(grid)):
        last = ""
        for c in range(len(grid[r])):
            if grid[r][c] == "":
                grid[r][c] = last
            last = grid[r][c]

    for r in range(1, len(grid)):
        for c in range(len(grid[r])):
            if grid[r][c] == "":
                grid[r][c] = grid[r - 1][c]

    return grid


def _headers_to_single_row(headers_grid: List[List[str]]) -> List[str]:
    """将多级表头压平成单行列名（备用方案）"""
    levels = len(headers_grid)
    cols = len(headers_grid[0]) if headers_grid else 0
    single: List[str] = []
    for c in range(cols):
        parts = []
        for r in range(levels):
            val = headers_grid[r][c].strip()
            if val:
                parts.append(val)
        single.append(" / ".join(parts) if parts else f"col_{c+1}")
    return single


